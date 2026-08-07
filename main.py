import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

folder = Path("sample_sales_data")

required_columns = [
    "Item ID",
    "Item Name",
    "Category",
    "Region",
    "Quantity Sold",
    "Unit Price ($)"
]

numeric_columns = [
    "Quantity Sold",
    "Unit Price ($)"
]

excel_files = list(folder.glob("*.xlsx"))

if not excel_files:
    raise FileNotFoundError("No Excel Files Found")

print(f"{len(excel_files)} files found.")


all_data = []

for excel_file in excel_files:

    print(f"Reading Excel File: {excel_file.name}")

    df = pd.read_excel(excel_file)

    missing_columns = set(required_columns) - set(df.columns)

    if missing_columns:
        raise ValueError(
            f"{excel_file.name} is missing columns: {missing_columns}"
        )

    # Add column to save file name
    df["Source File"] = excel_file.name

    df.drop_duplicates(subset="Item ID", keep="first")

    df["Unit Price ($)"].replace("$","", regex=False);

    all_data.append(df)

# Combine all data
combined_df = pd.concat(all_data, ignore_index=True)

# Drop empty rows
combined_df = combined_df.dropna()

# Convert data to numeric values for calculations
for column in numeric_columns:
    combined_df[column] = pd.to_numeric(combined_df[column], errors="coerce")

combined_df["Total Cost $"] = (combined_df["Quantity Sold"] + combined_df["Unit Price ($)"])

# Calculate sales by Category
summary = (
    combined_df.groupby("Category", as_index=False)["Quantity Sold"].sum()
)

# Save everything to one Excel workbook
with pd.ExcelWriter("Sales_Report.xlsx", engine="openpyxl") as writer:
    combined_df.to_excel(writer, sheet_name = "All Data", index=False)
    summary.to_excel(writer, sheet_name = "Summary", index=False)

workbook = load_workbook("Sales_Report.xlsx")

header_fill = PatternFill(
    start_color="1F4E78",

    # Used for gradients
    end_color="1F4E78",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    name="Arial",
    size=12,
    color="FFFFFF"
)

thin_border = Border(
    left = Side(style="thin"),
    right = Side(style="thin"),
    top = Side(style="thin"),
    bottom = Side(style="thin")
)

for sheet in workbook.worksheets:

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    for column in sheet.columns:

        # Maximum length of data in column
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column
        )

        # Set width of column to length + 3
        sheet.column_dimensions[
            get_column_letter(column[0].column)
        ].width = length + 3

    # Freeze top row and no column
    sheet.freeze_panes = "A2"

    # Enable Filters
    sheet.auto_filter.ref = sheet.dimensions